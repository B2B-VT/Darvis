from __future__ import annotations

import argparse
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any


DEFAULT_WORKBOOK = Path("/Users/kushpatel/Downloads/rag_llm_fallback_qa_test_workbook (1).xlsx")


@dataclass
class QATestCase:
    id: str
    category: str
    question: str
    expected: str
    risk_level: str
    expected_model_path: str
    data_dependency: str
    what_this_tests: str = ""
    row: dict[str, Any] | None = None


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _load_with_openpyxl(path: Path) -> list[dict[str, Any]] | None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Test Cases"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    return [
        {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        for row in rows[1:]
        if any(cell is not None for cell in row)
    ]


def _xlsx_text(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        v = cell.find("x:v", ns)
        return shared_strings[int(v.text)] if v is not None and v.text is not None else None
    if cell_type == "inlineStr":
        texts = [t.text or "" for t in cell.findall(".//x:t", ns)]
        return "".join(texts) or None
    v = cell.find("x:v", ns)
    return v.text if v is not None else None


def _column_index(cell_ref: str) -> int:
    letters = re.match(r"[A-Z]+", cell_ref.upper())
    if not letters:
        return 0
    value = 0
    for ch in letters.group(0):
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value - 1


def _load_with_stdlib(path: Path) -> list[dict[str, Any]]:
    ns = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(path) as zf:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("x:si", ns):
                shared_strings.append("".join(t.text or "" for t in si.findall(".//x:t", ns)))

        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root.findall("rel:Relationship", ns)}
        target = None
        for sheet in wb_root.findall("x:sheets/x:sheet", ns):
            if sheet.attrib.get("name") == "Test Cases":
                target = rels[sheet.attrib[f"{{{ns['r']}}}id"]]
                break
        if not target:
            raise ValueError("Workbook is missing a 'Test Cases' sheet.")
        target = target.lstrip("/")
        sheet_path = target if target.startswith("xl/") else "xl/" + target
        sheet_root = ET.fromstring(zf.read(sheet_path))

        parsed_rows: list[list[Any]] = []
        for row in sheet_root.findall(".//x:sheetData/x:row", ns):
            values: list[Any] = []
            for cell in row.findall("x:c", ns):
                idx = _column_index(cell.attrib.get("r", "A"))
                while len(values) <= idx:
                    values.append(None)
                values[idx] = _xlsx_text(cell, shared_strings, ns)
            parsed_rows.append(values)

    if not parsed_rows:
        return []
    headers = [_norm_header(h) for h in parsed_rows[0]]
    return [
        {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        for row in parsed_rows[1:]
        if any(cell is not None for cell in row)
    ]


def load_test_cases(path: str | Path = DEFAULT_WORKBOOK) -> list[QATestCase]:
    workbook = Path(path).expanduser()
    raw_rows = _load_with_openpyxl(workbook)
    if raw_rows is None:
        raw_rows = _load_with_stdlib(workbook)

    cases: list[QATestCase] = []
    for row in raw_rows:
        case_id = str(row.get("id") or "").strip()
        question = str(row.get("test_question") or "").strip()
        if not case_id or not question:
            continue
        cases.append(
            QATestCase(
                id=case_id,
                category=str(row.get("category") or "").strip(),
                question=question,
                expected=str(row.get("passing_criteria_expected_response") or "").strip(),
                risk_level=str(row.get("risk_level") or "").strip(),
                expected_model_path=str(row.get("expected_model_path") or "").strip(),
                data_dependency=str(row.get("data_dependency") or "").strip(),
                what_this_tests=str(row.get("what_this_tests") or "").strip(),
                row=row,
            )
        )
    return cases


def main() -> None:
    parser = argparse.ArgumentParser(description="Load Darvis RAG QA workbook cases.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    cases = load_test_cases(args.workbook)
    if args.json:
        print(json.dumps([asdict(c) for c in cases], indent=2))
    else:
        print(f"Loaded {len(cases)} test cases from {args.workbook}")
        for case in cases[:10]:
            print(f"{case.id}\t{case.category}\t{case.question}")


if __name__ == "__main__":
    main()
